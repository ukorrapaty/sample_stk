"""
Daily invest report generator
Assumptions (explicit):
 - "Original Price" is the first available price for the given trading day (first timestamp).
 - Invested value thresholds (lowerLimit / upperLimit) are percentages of the invested value (i.e., portfolio value = shares * price).
 - If intraday data for a given day is missing, the script will treat the last available price on that day as EndOfDay; if none exists, the entry is recorded but numeric fields may be NaN.
"""

import os
from datetime import datetime, timedelta
from collections import defaultdict
import pandas as pd
import stocks as stk
from tqdm import tqdm
from openpyxl import load_workbook
import stock_data as sd
import stock_simulate_trade as simtr
import reports as rpt


# -------------------------
# User-provided configuration
# -------------------------
StocksFile = r"C:/Users/ukorr/OneDrive/Documents/Projects/work/stockmarket/TrackStocks/src/input/stocks.xlsx"
#StocksFile = r"C:/Users/ukorr/OneDrive/Documents/Projects/work/stockmarket/TrackStocks/src/input/stocks_s_p.xlsx"
NumberOfDays = 60 #8 #60
DailyInitialInvestment = 50000.0  # $10000
OutputDir = r"C:/Users/ukorr/OneDrive/Documents/Projects/work/stockmarket/TrackStocks/DailyInvestReport"

# -------------------------

# Derived / fixed
# -------------------------
os.makedirs(OutputDir, exist_ok=True)
timestamp = datetime.now().strftime("%Y%m%d%H%M")
output_excel_file = os.path.join(OutputDir, f"daily_invest_report_{timestamp}.xlsx")
details_excel_file = os.path.join(OutputDir, f"details_daily_invest_report_{timestamp}.xlsx")
last30DaySheet="DetailsPast30Days"
monthlyReportSheet="MonthlyReport"
investInfoSheet="OverallInvestmentInfo"
summarySheet="Summary"

def save_data(overall_info, details_df, summary_df):
    # Duplicates df
    #duplicates_df = pd.DataFrame({"DuplicateSymbol": duplicates}) if duplicates else pd.DataFrame(columns=["DuplicateSymbol"])

    # Save to Excel with multiple sheets
    print(f"Writing output to Excel: {output_excel_file}")
    with pd.ExcelWriter(output_excel_file, engine="openpyxl") as writer:
        details_df.to_excel(writer, sheet_name=last30DaySheet, index=False)
        summary_df.to_excel(writer, sheet_name=summarySheet, index=False)
        # overall info as a small df
        overall_df = pd.DataFrame([overall_info])
        overall_df.to_excel(writer, sheet_name=investInfoSheet, index=False)
        #duplicates_df.to_excel(writer, sheet_name="Duplicates", index=False)

def convert_datetimes_to_naive(df):
    for col_name, dtype in df.dtypes.items():
        if pd.api.types.is_datetime64_any_dtype(dtype):
            # Convert to UTC and remove timezone information (tz_localize(None))
            df[col_name] = df[col_name].dt.tz_convert('UTC').dt.tz_localize(None)
    return df

def save_stock_data(stocks,symbol_data):
    # Save to Excel with multiple sheets
    print(f"Writing output to Excel: {details_excel_file}")
    
    for index, sym_row in stocks.iterrows():
        sym=sym_row["stock"]
        df_stock = symbol_data.get(sym, pd.DataFrame(columns=["Symbol","Name","Datetime","Price"]))
        df_stock=convert_datetimes_to_naive(df_stock)
        if not os.path.exists(details_excel_file):
            df_stock.to_excel(details_excel_file, sheet_name="Details", index=False)
        else:
            # Load the workbook and set up the writer
            book = load_workbook(details_excel_file)
            writer = pd.ExcelWriter(details_excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    
            # Keep the existing sheets loaded into the writer
            writer._book = book
            writer._sheets = {ws.title: ws for ws in book.worksheets}

            # Determine the next empty row
            # Use max_row of the specific sheet
            startrow = writer._sheets["Details"].max_row
    
            # Append the new DataFrame
            # Set header=False to avoid writing the column names again
            # Set index=False to avoid writing the pandas index
            df_stock.to_excel(writer, sheet_name="Details", startrow=startrow, index=False, header=False)
            # overall info as a small df
            #duplicates_df.to_excel(writer, sheet_name="Duplicates", index=False)
            writer.close()



# -------------------------
# Main process
# -------------------------
def main():
    print("Reading stocks file...")
    #stocks, duplicates = stk.read_stocks_file(StocksFile)
    stocks=stk.process_stock_file(StocksFile)
    #print(f"Unique symbols: {len(stocks)}; duplicates found: {len(duplicates)}")

    if stocks.empty:
        print("No stocks found. Exiting.")
        return

    print("Fetching intraday 30-minute data for each symbol...")
    symbol_data_map = sd.fetch_intraday_30min(stocks, period_days=NumberOfDays)
    #save_stock_data(stocks,symbol_data_map)

    txn_data = simtr.get_txn_data(stocks, DailyInitialInvestment, NumberOfDays, symbol_data_map)
    
    details_df = rpt.get_details_df(txn_data)

    summary_rows, totalInvestedAll, totalSoldAll = rpt.get_summary_data(details_df)

    total_gain_loss_all = totalSoldAll - totalInvestedAll
    total_gain_loss_pct_all = (total_gain_loss_all / totalInvestedAll * 100) if totalInvestedAll != 0 else 0.0

    summary_df = pd.DataFrame(summary_rows, columns=["Symbol","Name","TotalInvested","TotalSoldValue","TotalGainLoss","TotalGainLossPercent"])

    overall_info = {
        "TotalInvestedAll": round(totalInvestedAll, 2),
        "TotalSoldAll": round(totalSoldAll, 2),
        "TotalGainLossAll": round(total_gain_loss_all, 2),
        "TotalGainLossPercentAll": round(total_gain_loss_pct_all, 2)
    }

    save_data(overall_info, details_df, summary_df)

    print("Done.")
    rpt.generate_monthly_report(output_excel_file, last30DaySheet, monthlyReportSheet)
    print(f"Report saved to: {output_excel_file}")

if __name__ == "__main__":
    main()
